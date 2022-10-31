import openpyxl

#criar planilha
book = openpyxl.Workbook()
# Visualizar pagina existente
print(book.sheetnames)
# Criar uma pagina
book.create_sheet('Frutas')
# como selecionar uma pagina
frutas_page = book['Frutas']
frutas_page.append(['Frutas', 'Quantidade','Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Fruta 2', '2', 'R$15,90'])
frutas_page.append(['Fruta 3', '10', 'R$30,90'])
frutas_page.append(['Fruta 4', '2', 'R$50,50'])
# Salvar Planilha
book.save('Planilha de compras.xlsx')
